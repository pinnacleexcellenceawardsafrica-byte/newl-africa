import os
import platform
import copy as copy_module
import threading
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (
    OneCellAnchor, TwoCellAnchor, AnchorMarker, AbsoluteAnchor
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels, DEFAULT_COLUMN_WIDTH, DEFAULT_ROW_HEIGHT
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
from .models import Site, PurchaseOrder
from openpyxl.utils import get_column_letter
import shutil
import subprocess
import re
import time
import traceback

IS_WINDOWS = platform.system() == 'Windows'

# ============================================================
# EXCEL COM SERIALIZATION LOCK
# ============================================================
EXCEL_COM_LOCK = threading.Lock()

RPC_E_CALL_REJECTED = -2147418111


def _is_call_rejected_error(exc):
    """Detect the RPC_E_CALL_REJECTED signature."""
    msg = str(exc)
    return (
        'RPC_E_CALL_REJECTED' in msg or
        '-2147418111' in msg or
        '0x80010001' in msg or
        'was rejected by callee' in msg.lower()
    )


_LIBREOFFICE_AVAILABLE = None


def check_libreoffice_available(force_recheck=False):
    """Check once (and cache) whether LibreOffice ('soffice') is on PATH."""
    global _LIBREOFFICE_AVAILABLE
    if _LIBREOFFICE_AVAILABLE is not None and not force_recheck:
        return _LIBREOFFICE_AVAILABLE

    common_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice\program\soffice.bin",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.bin",
        r"C:\Program Files\LibreOffice\program\soffice",
        r"C:\Program Files (x86)\LibreOffice\program\soffice",
    ]

    soffice_path = shutil.which('soffice')

    if not soffice_path:
        for path in common_paths:
            if os.path.exists(path):
                soffice_path = path
                break

    if not soffice_path and IS_WINDOWS:
        try:
            program_files = os.environ.get('ProgramFiles', 'C:\\Program Files')
            program_files_x86 = os.environ.get('ProgramFiles(x86)', 'C:\\Program Files (x86)')

            for base_dir in [program_files, program_files_x86]:
                libre_dir = os.path.join(base_dir, 'LibreOffice')
                if os.path.exists(libre_dir):
                    for item in os.listdir(libre_dir):
                        program_dir = os.path.join(libre_dir, item, 'program')
                        if os.path.exists(program_dir):
                            soffice_exe = os.path.join(program_dir, 'soffice.exe')
                            if os.path.exists(soffice_exe):
                                soffice_path = soffice_exe
                                break
                    if soffice_path:
                        break
        except Exception as e:
            print(f"⚠️ Could not search for LibreOffice: {e}")

    _LIBREOFFICE_AVAILABLE = soffice_path is not None

    if _LIBREOFFICE_AVAILABLE:
        print(f"✅ LibreOffice found: {soffice_path}")
    else:
        print(
            "⚠️ LibreOffice ('soffice') is NOT installed on this machine.\n"
            "   PDF conversion will fall back to unattended Excel COM automation,\n"
            "   which is unsupported by Microsoft for server-side use and is prone\n"
            "   to RPC_E_CALL_REJECTED and 'Workbooks.Open' failures.\n"
            "   Fix: winget install TheDocumentFoundation.LibreOffice  (Windows)\n"
            "        apt-get install libreoffice                       (Linux)"
        )

    return _LIBREOFFICE_AVAILABLE


def _find_soffice():
    """Locate the soffice executable."""
    soffice_cmd = shutil.which('soffice')
    if soffice_cmd:
        return soffice_cmd

    common_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for path in common_paths:
        if os.path.exists(path):
            return path

    try:
        program_files = os.environ.get('ProgramFiles', 'C:\\Program Files')
        program_files_x86 = os.environ.get('ProgramFiles(x86)', 'C:\\Program Files (x86)')

        for base_dir in [program_files, program_files_x86]:
            libre_dir = os.path.join(base_dir, 'LibreOffice')
            if os.path.exists(libre_dir):
                for item in os.listdir(libre_dir):
                    program_dir = os.path.join(libre_dir, item, 'program')
                    if os.path.exists(program_dir):
                        soffice_exe = os.path.join(program_dir, 'soffice.exe')
                        if os.path.exists(soffice_exe):
                            return soffice_exe
    except Exception as e:
        print(f"⚠️ Could not search for LibreOffice: {e}")

    return None


def ensure_excel_com_desktop_folder():
    """Ensure the Desktop folder exists."""
    if not IS_WINDOWS:
        return
    try:
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        if not os.path.exists(desktop_path):
            os.makedirs(desktop_path, exist_ok=True)
            print(f"🔧 Created missing Desktop folder for Excel COM: {desktop_path}")
    except Exception as e:
        print(f"⚠️ Could not verify/create Desktop folder: {e}")


# ============================================================
# SITE MODULE PACKAGE → TEMPLATE SHEET MAPPING
# ============================================================
SITE_MODULE_TO_SHEET = {
    'TSS LOS': 'TSS LOS',
    'TI New Phase': 'RAN',
    'TI MW Phase': 'MWLINK',
    'TSS Workflow': 'TSS',
    'TI MW Dismantling & Warehousing': 'MWLINK',
    'Relocation': 'MWLINK',
    'TI Dismantling and Warehousing': 'MWLINK',
    'TI Dismantling & Warehousing': 'MWLINK',
    'TI Modernization 2G3G LTE': 'Modernization',
    'TI Cabinet Installation': 'Cabinet',
    'TI New Phase-Transport': 'RAN',
    'KE Clean-up Workflow': 'RAN',
    'KE Clean-UP Workflow': 'RAN',
    'TI MW Phase-Transport': 'TSS LOS',
    'TI MW dismantling and relocation': 'MWLINK',
    'TI MW dismantling & relocation': 'MWLINK',
}

SHEET_TO_PM = {
    'MWLINK': 'Ken Juma',
    'TSS LOS': 'Ken Juma',
    'Modernization': 'Julius Kamemba',
    'RAN': 'Julius Kamemba',
    'TSS': 'Julius Kamemba',
    'Cabinet': 'Julius Kamemba',
}


def get_sheet_from_module(site_module_package):
    """Get template sheet name from SITE MODULE PACKAGE"""
    if not site_module_package:
        return 'Modernization'

    module = site_module_package.strip()

    if module in SITE_MODULE_TO_SHEET:
        return SITE_MODULE_TO_SHEET[module]

    for key, sheet in SITE_MODULE_TO_SHEET.items():
        if key.lower() in module.lower() or module.lower() in key.lower():
            return sheet

    print(f"⚠️ No sheet mapping found for: {site_module_package}, using default 'Modernization'")
    return 'Modernization'


def get_pm_from_sheet(sheet_name):
    """Get Project Manager from the sheet name."""
    if not sheet_name:
        return 'Julius Kamemba'

    sheet = sheet_name.strip()

    if sheet in SHEET_TO_PM:
        return SHEET_TO_PM[sheet]

    for key, pm in SHEET_TO_PM.items():
        if key.lower() in sheet.lower() or sheet.lower() in key.lower():
            return pm

    print(f"⚠️ No PM mapping found for sheet: {sheet_name}, using default 'Julius Kamemba'")
    return 'Julius Kamemba'


def is_double_site(site_id):
    """Check if a site ID has two site names (contains '-')"""
    if not site_id:
        return False
    parts = site_id.split('-')
    if len(parts) >= 2:
        import re
        site_pattern = re.compile(r'^[A-Z]{2,4}\d{3,4}$')
        first_part = parts[0]
        second_part = parts[1]
        if site_pattern.match(first_part) and site_pattern.match(second_part):
            return True
        if len(parts) >= 2:
            return True
    return False


def get_pm_for_site(site_module_package, site_id, site_name=None):
    """
    Get the correct PM for a site based on module package and site ID.
    
    Special rules:
    - TI Dismantling and Warehousing: double site IDs → Ken Juma, single → Julius Kamemba
    - TI MW Dismantling & Warehousing: double site IDs → Ken Juma, single → Julius Kamemba
    - TI MW dismantling and relocation: Ken Juma
    - Relocation: Julius Kamemba (per your PM sheet)
    """
    if not site_module_package:
        return 'Julius Kamemba'

    module = site_module_package.strip().lower()

    # TI MW dismantling and relocation → Ken Juma
    if 'ti mw dismantling' in module and 'relocation' in module:
        return 'Ken Juma'
    
    # Relocation → Julius Kamemba (per your PM sheet)
    if module == 'relocation' or module.startswith('relocation'):
        return 'Julius Kamemba'

    # TI Dismantling and Warehousing - check double site rule
    if 'dismantling' in module and ('warehousing' in module or 'warehouse' in module):
        if is_double_site(site_id):
            print(f"   🔍 Double site detected: {site_id} → Ken Juma")
            return 'Ken Juma'
        else:
            print(f"   🔍 Single site detected: {site_id} → Julius Kamemba")
            return 'Julius Kamemba'

    # Get sheet and default PM
    sheet = get_sheet_from_module(site_module_package)
    return get_pm_from_sheet(sheet)


def get_region_from_id(site_id):
    """Extract region from site ID prefix."""
    region_map = {
        'BIA': 'Busia', 'VGA': 'Vihiga', 'BMA': 'Bungoma',
        'KGA': 'Kirinyaga', 'TKA': 'Thika', 'NYI': 'Nyeri',
        'MEC': 'Meru', 'NRU': 'Nakuru', 'HBY': 'Homa Bay',
        'MUA': 'Machakos', 'MOI': 'Uasin Gishu', 'TNZ': 'Trans Nzoia',
        'BGO': 'Baringo', 'WPT': 'West Pokot', 'MEN': 'Meru',
        'TSO': 'Trans Nzoia', 'MOS': 'Machakos', 'NBI': 'Nairobi',
        'MGA': 'Machakos', 'NUA': 'Nakuru', 'SRU': 'Samburu',
        'THE': 'Tharaka Nithi', 'MES': 'Meru'
    }
    prefix = site_id.split('-')[0] if '-' in site_id else site_id[:3]
    return region_map.get(prefix, '')


def is_merged_cell(worksheet, row, col):
    """Check if a cell is merged"""
    for merged_range in worksheet.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and \
           col in range(merged_range.min_col, merged_range.max_col + 1):
            return True
    return False


def is_merged_write_blocked(worksheet, row, col):
    """Check if writing to this cell is unsafe because it is part of a merged range but NOT the top-left anchor"""
    for merged_range in worksheet.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and \
           col in range(merged_range.min_col, merged_range.max_col + 1):
            is_anchor = (row == merged_range.min_row and col == merged_range.min_col)
            if not is_anchor:
                return True
    return False


def get_merged_anchor(worksheet, row, col):
    """
    Get the anchor (top-left) cell of a merged range that contains (row, col).
    Returns (anchor_row, anchor_col) or (row, col) if not merged.
    """
    for merged_range in worksheet.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and \
           col in range(merged_range.min_col, merged_range.max_col + 1):
            return merged_range.min_row, merged_range.min_col
    return row, col


def kill_excel_processes():
    """Force kill any hanging Excel/LibreOffice processes."""
    try:
        if IS_WINDOWS:
            print("🔄 Killing hanging Excel processes (Windows)...")
            subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                          capture_output=True, timeout=5)
        else:
            print("🔄 Killing hanging LibreOffice processes (Linux)...")
            subprocess.run(['pkill', '-f', 'soffice'],
                          capture_output=True, timeout=5)
        time.sleep(1.5)
        return True
    except Exception as e:
        print(f"⚠️ Could not kill office processes: {e}")
        return False


def reset_win32com_gen_py_cache():
    """Delete the cached win32com Excel type-library wrapper."""
    if not IS_WINDOWS:
        return False

    try:
        import pythoncom
        import win32com
        from win32com.client import gencache

        cache_dir = getattr(win32com, '__gen_path__', None)
        if not cache_dir:
            cache_dir = os.path.join(os.environ.get('TEMP', ''), 'gen_py')

        if cache_dir and os.path.exists(cache_dir):
            print(f"🔧 Clearing stale win32com gen_py cache: {cache_dir}")
            shutil.rmtree(cache_dir, ignore_errors=True)

        pythoncom.CoInitialize()
        try:
            gencache.EnsureDispatch('Excel.Application').Quit()
            print("✅ win32com gen_py cache rebuilt")
            return True
        finally:
            pythoncom.CoUninitialize()

    except Exception as e:
        print(f"⚠️ Could not rebuild win32com gen_py cache: {e}")
        return False


def _is_gen_py_cache_error(exc):
    """Detect the specific 'stale gen_py cache' AttributeError signature."""
    msg = str(exc)
    return 'gen_py' in msg and (
        'CLSIDToClassMap' in msg or 'CLSIDToPackageMap' in msg
    )


def safe_delete_file(file_path):
    """Safely delete a file with aggressive retry logic"""
    if not os.path.exists(file_path):
        return True

    kill_excel_processes()

    for attempt in range(5):
        try:
            os.remove(file_path)
            print(f"✅ Deleted: {file_path}")
            return True
        except PermissionError:
            print(f"⚠️ Permission denied (attempt {attempt+1}/5), waiting...")
            time.sleep(2)
            kill_excel_processes()
        except Exception as e:
            print(f"⚠️ Delete failed: {e}")
            time.sleep(1)

    try:
        temp_path = file_path + '.old'
        os.rename(file_path, temp_path)
        time.sleep(0.5)
        os.remove(temp_path)
        print(f"✅ Deleted via rename: {file_path}")
        return True
    except:
        pass

    return False


def safe_save_workbook(wb, output_path, max_retries=5):
    """Save workbook robustly using atomic write pattern."""
    kill_excel_processes()
    time.sleep(1)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    temp_path = f"{output_path}.{os.getpid()}.{int(time.time() * 1000)}.tmp"

    saved = False
    for attempt in range(max_retries):
        try:
            kill_excel_processes()
            time.sleep(0.5)
            wb.save(temp_path)
            saved = True
            break
        except PermissionError:
            print(f"⚠️ Permission denied writing temp (attempt {attempt+1}/{max_retries})")
            kill_excel_processes()
            time.sleep(2)
        except Exception as e:
            print(f"⚠️ Save failed: {e}")
            time.sleep(1)

    if not saved:
        print(f"❌ Could not write workbook to temp file: {output_path}")
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass
        return False

    try:
        wb.close()
    except:
        pass

    time.sleep(0.5)

    if os.path.exists(output_path):
        for attempt in range(max_retries):
            try:
                kill_excel_processes()
                time.sleep(0.5)
                os.remove(output_path)
                print(f"✅ Removed existing target: {output_path}")
                break
            except PermissionError:
                print(f"⚠️ Permission denied removing target (attempt {attempt+1}/{max_retries})")
                kill_excel_processes()
                time.sleep(2)
            except Exception as e:
                print(f"⚠️ Could not remove target: {e}")
                time.sleep(1)

    for attempt in range(max_retries):
        try:
            kill_excel_processes()
            time.sleep(0.5)
            shutil.move(temp_path, output_path)
            print(f"✅ Saved: {output_path}")
            return True
        except PermissionError:
            print(f"⚠️ Permission denied moving (attempt {attempt+1}/{max_retries})")
            kill_excel_processes()
            time.sleep(2)
        except Exception as e:
            print(f"⚠️ Move failed: {e}")
            time.sleep(1)

    try:
        if os.path.exists(temp_path):
            os.remove(temp_path)
    except:
        pass

    return False


# ============================================================
# IMAGE (LOGO) SIZE/POSITION STABILITY
# ============================================================

def _excel_col_width_to_px(width):
    """Convert an Excel column width (character units) to pixels."""
    if width is None:
        width = DEFAULT_COLUMN_WIDTH
    return int(round(width * 7 + 5))


def _col_width_px(ws, col_idx):
    """Pixel width of column `col_idx` (1-based) on worksheet `ws`."""
    col_letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else None
    if width is None:
        width = getattr(getattr(ws, 'sheet_format', None), 'defaultColWidth', None)
    return _excel_col_width_to_px(width)


def _row_height_px(ws, row_idx):
    """Pixel height of row `row_idx` (1-based) on worksheet `ws`."""
    dim = ws.row_dimensions.get(row_idx)
    height_pt = dim.height if (dim and dim.height) else None
    if height_pt is None:
        height_pt = getattr(getattr(ws, 'sheet_format', None), 'defaultRowHeight', None)
    if height_pt is None:
        height_pt = DEFAULT_ROW_HEIGHT
    return points_to_pixels(height_pt)


def _marker_x_px(ws, col_idx_zero_based, col_off_emu):
    """Absolute X pixel position of an anchor marker."""
    from openpyxl.utils.units import EMU_to_pixels
    x_px = sum(_col_width_px(ws, c + 1) for c in range(col_idx_zero_based))
    return x_px + EMU_to_pixels(col_off_emu)


def _marker_y_px(ws, row_idx_zero_based, row_off_emu):
    """Absolute Y pixel position of an anchor marker."""
    from openpyxl.utils.units import EMU_to_pixels
    y_px = sum(_row_height_px(ws, r + 1) for r in range(row_idx_zero_based))
    return y_px + EMU_to_pixels(row_off_emu)


def _fixed_size_anchor_for(source_sheet, anchor, fallback_width_px=None, fallback_height_px=None):
    """Return a new OneCellAnchor with fixed size."""
    try:
        if isinstance(anchor, TwoCellAnchor):
            frm, to = anchor._from, anchor.to
            x0 = _marker_x_px(source_sheet, frm.col, frm.colOff)
            y0 = _marker_y_px(source_sheet, frm.row, frm.rowOff)
            x1 = _marker_x_px(source_sheet, to.col, to.colOff)
            y1 = _marker_y_px(source_sheet, to.row, to.rowOff)
            width_px = max(1, x1 - x0)
            height_px = max(1, y1 - y0)
            new_from = AnchorMarker(col=frm.col, colOff=frm.colOff, row=frm.row, rowOff=frm.rowOff)

        elif isinstance(anchor, OneCellAnchor):
            frm = anchor._from
            if anchor.ext and anchor.ext.cx and anchor.ext.cy:
                return None
            if fallback_width_px is None or fallback_height_px is None:
                return None
            width_px, height_px = fallback_width_px, fallback_height_px
            new_from = AnchorMarker(col=frm.col, colOff=frm.colOff, row=frm.row, rowOff=frm.rowOff)

        else:
            return None

        ext = XDRPositiveSize2D(cx=pixels_to_EMU(width_px), cy=pixels_to_EMU(height_px))
        return OneCellAnchor(_from=new_from, ext=ext)

    except Exception as e:
        print(f"⚠️ Could not compute fixed-size anchor for image, keeping original anchor: {e}")
        return None


def copy_sheet_with_openpyxl(template_path, sheet_name, output_path):
    """Copy sheet using openpyxl, including embedded images."""
    try:
        print(f"📋 Copying sheet '{sheet_name}' using openpyxl...")

        kill_excel_processes()
        time.sleep(1)

        wb = openpyxl.load_workbook(template_path)

        actual_sheet_name = None
        if sheet_name in wb.sheetnames:
            actual_sheet_name = sheet_name
        else:
            for name in wb.sheetnames:
                if name.strip().lower() == sheet_name.strip().lower():
                    actual_sheet_name = name
                    break

        if not actual_sheet_name:
            actual_sheet_name = wb.sheetnames[0]
            print(f"⚠️ Sheet not found, using: {actual_sheet_name}")

        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)

        source_sheet = wb[actual_sheet_name]
        new_sheet = new_wb.create_sheet(title=actual_sheet_name)

        # Copy cell values and styles
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                if cell.value is not None:
                    new_cell.value = cell.value
                if cell.has_style:
                    try:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = cell.alignment.copy()
                    except:
                        pass

        # Copy default column width
        try:
            if source_sheet.sheet_format and source_sheet.sheet_format.defaultColWidth:
                new_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
        except Exception as e:
            print(f"⚠️ Could not copy default column width: {e}")

        # Copy default row height
        try:
            if source_sheet.sheet_format and source_sheet.sheet_format.defaultRowHeight:
                new_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
        except Exception as e:
            print(f"⚠️ Could not copy default row height: {e}")

        # Copy column dimensions
        for col, dim in source_sheet.column_dimensions.items():
            new_dim = new_sheet.column_dimensions[col]
            if dim.width:
                new_dim.width = dim.width
            new_dim.hidden = dim.hidden

        # Copy row dimensions
        for row in source_sheet.row_dimensions:
            if source_sheet.row_dimensions[row].height:
                new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(
                start_row=merged_range.min_row,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row,
                end_column=merged_range.max_col
            )

        # Copy print area / page setup
        try:
            if source_sheet.print_area:
                new_sheet.print_area = source_sheet.print_area
            new_sheet.page_setup.orientation = source_sheet.page_setup.orientation or 'landscape'
            new_sheet.page_margins = copy_module.copy(source_sheet.page_margins)
        except Exception as e:
            print(f"⚠️ Could not copy print area/page setup: {e}")

        # Copy images with fixed sizing
        copied_images = 0
        for img in getattr(source_sheet, '_images', []):
            try:
                img_bytes = img._data()
                new_img = XLImage(BytesIO(img_bytes))

                fixed_anchor = _fixed_size_anchor_for(
                    source_sheet, img.anchor,
                    fallback_width_px=getattr(new_img, 'width', None),
                    fallback_height_px=getattr(new_img, 'height', None),
                )
                new_img.anchor = fixed_anchor if fixed_anchor is not None else copy_module.deepcopy(img.anchor)

                new_sheet.add_image(new_img)
                copied_images += 1
            except Exception as e:
                print(f"⚠️ Could not copy an image: {e}")
        print(f"🖼️ Copied {copied_images} embedded image(s) with fixed sizing")

        try:
            new_sheet.sheet_state = 'visible'
        except Exception:
            pass

        try:
            wb.close()
        except:
            pass

        success = safe_save_workbook(new_wb, output_path)

        if success:
            print(f"✅ Single-sheet workbook created: {output_path}")
            return True
        else:
            print(f"❌ Failed to save workbook: {output_path}")
            return False

    except Exception as e:
        print(f"❌ openpyxl copy failed: {e}")
        traceback.print_exc()
        return False


# ============================================================
# ROBUST HEADER MATCHING
# ============================================================
def _find_header_col(headers, candidates):
    """Return (column_index, matched_header_text) for the first candidate."""
    for c in candidates:
        if c in headers:
            return headers[c], c

    lower_headers = {k.lower().strip(): (k, v) for k, v in headers.items()}
    for c in candidates:
        hit = lower_headers.get(c.lower().strip())
        if hit:
            return hit[1], hit[0]

    for c in candidates:
        c_lower = c.lower().strip()
        for header_text, col in headers.items():
            header_lower = header_text.lower().strip()
            if c_lower in header_lower or header_lower in c_lower:
                return col, header_text

    return None, None


def process_book2(file_path, uploaded_file):
    """Process Book2.xlsx - Site master data"""
    print(f"📂 Processing Book2: {file_path}")

    kill_excel_processes()
    time.sleep(0.5)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    print(f"📋 Headers found: {list(headers.keys())}")

    col_site_id, hdr_site_id = _find_header_col(headers, ['Site ID', 'SITE ID', 'Site_ID', 'SiteID'])
    col_site_name, hdr_site_name = _find_header_col(headers, ['Site Name', 'SITE NAME', 'Site_Name', 'Site'])
    col_smp, hdr_smp = _find_header_col(headers, ['Smp', 'SMP', 'SMP ID', 'SMP_ID'])
    col_module, hdr_module = _find_header_col(headers, ['Site Module Package', 'SITE MODULE PACKAGE', 'Module Package'])
    col_po_number, hdr_po_number = _find_header_col(headers, ['sPO Number', 'PO Number', 'PO_Number', 'SPO Number'])
    col_po_date, hdr_po_date = _find_header_col(headers, ['SPO Date', 'PO Date', 'PO_Date'])
    col_pm, hdr_pm = _find_header_col(headers, ['Nokia Project Manager', 'Nokia PM', 'Project Manager'])
    col_sub_category, hdr_sub_category = _find_header_col(headers, ['SUB CATEGORY', 'Sub Category', 'SubCategory'])
    col_project_category, hdr_project_category = _find_header_col(headers, ['Project Category', 'PROJECT CATEGORY'])
    col_no, hdr_no = _find_header_col(headers, ['No.', 'No', 'NO.', 'S/No', 'S/No.'])
    col_wcc_status, hdr_wcc_status = _find_header_col(headers, ['WCC Status', 'WCC_STATUS', 'Status'])
    col_region, hdr_region = _find_header_col(headers, ['Region', 'REGION'])

    print("📋 Resolved column mapping:")
    print(f"   Site ID          -> {hdr_site_id!r} (col {col_site_id})")
    print(f"   Site Name        -> {hdr_site_name!r} (col {col_site_name})")
    print(f"   Smp              -> {hdr_smp!r} (col {col_smp})")
    print(f"   Site Module Pkg  -> {hdr_module!r} (col {col_module})")
    print(f"   PO Number        -> {hdr_po_number!r} (col {col_po_number})")
    print(f"   PO Date          -> {hdr_po_date!r} (col {col_po_date})")
    print(f"   Nokia PM         -> {hdr_pm!r} (col {col_pm})")
    print(f"   Sub Category     -> {hdr_sub_category!r} (col {col_sub_category})")
    print(f"   Project Category -> {hdr_project_category!r} (col {col_project_category})")
    print(f"   No.              -> {hdr_no!r} (col {col_no})")
    print(f"   WCC Status       -> {hdr_wcc_status!r} (col {col_wcc_status})")
    print(f"   Region           -> {hdr_region!r} (col {col_region})")

    if not col_site_name:
        print("⚠️ No 'Site Name' column could be matched in Book2 — Site Name will fall back to Site ID.")

    site_count = 0

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        site_id = row[col_site_id - 1].value if col_site_id else None
        site_name = row[col_site_name - 1].value if col_site_name else None
        smp = row[col_smp - 1].value if col_smp else None
        site_module_package = row[col_module - 1].value if col_module else ''
        po_number = row[col_po_number - 1].value if col_po_number else None
        po_date = row[col_po_date - 1].value if col_po_date else None
        nokia_pm = row[col_pm - 1].value if col_pm else ''
        sub_category = row[col_sub_category - 1].value if col_sub_category else ''
        project_category = row[col_project_category - 1].value if col_project_category else ''
        no = row[col_no - 1].value if col_no else ''
        wcc_status = row[col_wcc_status - 1].value if col_wcc_status else None
        region_value = row[col_region - 1].value if col_region else None

        if not site_id or not smp or not po_number:
            print(f"⚠️ Skipping row {row_idx}: Missing SMP or PO Number")
            continue

        if not site_name:
            print(f"⚠️ Row {row_idx} (SMP {smp}): Site Name is blank — using Site ID '{site_id}' as fallback")

        if isinstance(po_date, datetime):
            po_date = po_date.strftime('%Y-%m-%d')
        elif po_date:
            po_date = str(po_date)

        # Get region from the Region column if available, otherwise from site ID
        region = str(region_value).strip() if region_value else get_region_from_id(str(site_id))
        if not region:
            region = get_region_from_id(str(site_id))
            print(f"⚠️ No region found in Region column for {site_id}, derived '{region}' from ID")

        # Get the correct PM using the site ID (for double/single site logic)
        pm_from_module = get_pm_for_site(
            str(site_module_package) if site_module_package else '',
            str(site_id) if site_id else None,
            str(site_name) if site_name else None
        )

        if pm_from_module:
            nokia_pm = pm_from_module

        sheet_name = get_sheet_from_module(str(site_module_package) if site_module_package else '')

        site, created = Site.objects.get_or_create(
            smp=str(smp).strip(),
            defaults={
                'site_id': str(site_id).strip(),
                'site_name': str(site_name).strip() if site_name else str(site_id).strip(),
                'site_module_package': str(site_module_package).strip() if site_module_package else '',
                'po_number': str(po_number).strip(),
                'po_date': po_date or '',
                'wcc_status': str(wcc_status).strip() if wcc_status else 'Pending',
                'nokia_pm': str(nokia_pm).strip() if nokia_pm else 'Julius Kamemba',
                'sub_category': str(sub_category).strip() if sub_category else '',
                'project_category': str(project_category).strip() if project_category else '',
                'no': str(no).strip() if no else '',
                'region': region,
                'source_file': uploaded_file
            }
        )

        if not created:
            site.site_id = str(site_id).strip()
            site.site_name = str(site_name).strip() if site_name else str(site_id).strip()
            site.site_module_package = str(site_module_package).strip() if site_module_package else ''
            site.po_number = str(po_number).strip()
            site.po_date = po_date or ''
            if wcc_status:
                site.wcc_status = str(wcc_status).strip()
            site.nokia_pm = str(nokia_pm).strip() if nokia_pm else 'Julius Kamemba'
            site.sub_category = str(sub_category).strip() if sub_category else ''
            site.project_category = str(project_category).strip() if project_category else ''
            site.region = region
            site.source_file = uploaded_file
            site.save()

        site_count += 1
        print(f"   ✅ Site {site_count}: SMP={smp} → PO={po_number} ({site.site_name}) [PM: {nokia_pm}] [Sheet: {sheet_name}] [Region: {region}]")

    uploaded_file.row_count = site_count
    uploaded_file.save()

    try:
        wb.close()
    except:
        pass

    print(f"✅ Book2 processed: {site_count} sites")
    return site_count


def process_purchase_orders(file_path, uploaded_file):
    """Process Purchase Orders.xlsx - PO line items."""
    print(f"📂 Processing Purchase Orders: {file_path}")

    kill_excel_processes()
    time.sleep(0.5)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    print(f"📋 Headers found: {list(headers.keys())}")

    po_count = 0
    sites_updated = set()
    unmatched_count = 0
    regions_corrected = 0

    po_col, _ = _find_header_col(headers, ['PO Number', 'PO_Number', 'sPO Number', 'SPO Number'])
    po_date_col, _ = _find_header_col(headers, ['PO Date', 'PO_Date', 'SPO Date'])
    item_col, item_col_name = _find_header_col(headers, ['Item No', 'Item No.', 'No.', 'S/No'])
    desc_col, desc_col_name = _find_header_col(headers, ['Item Description', 'Description', 'SID (Description)', 'SID'])
    qty_col, qty_col_name = _find_header_col(headers, ['Quantity', 'Qty'])
    unit_col, unit_col_name = _find_header_col(headers, ['Unit'])
    region_col, region_col_name = _find_header_col(headers, ['Region'])

    if not po_col:
        po_col = 1
    if not po_date_col:
        po_date_col = 2
    if not item_col:
        item_col = 3
        item_col_name = 'Item No'
    if not desc_col:
        desc_col = 6
        desc_col_name = 'Item Description'
    if not qty_col:
        qty_col = 5
        qty_col_name = 'Quantity'

    print(f"📋 Using columns:")
    print(f"   PO: {po_col}")
    print(f"   Item: {item_col_name} (column {item_col})")
    print(f"   Description: {desc_col_name} (column {desc_col})")
    print(f"   Qty: {qty_col_name} (column {qty_col})")
    if unit_col:
        print(f"   Unit: {unit_col_name} (column {unit_col})")
    if region_col:
        print(f"   Region: {region_col_name} (column {region_col})")

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        po_number = row[po_col - 1].value if po_col else None

        if not po_number:
            continue

        po_number_str = str(po_number).strip()

        sites = Site.objects.filter(po_number=po_number_str)

        if not sites.exists():
            try:
                po_int = int(float(po_number_str))
                sites = Site.objects.filter(po_number=str(po_int))
            except:
                pass

        if not sites.exists():
            unmatched_count += 1
            if unmatched_count <= 20:
                print(f"⚠️ No sites found for PO Number: {repr(po_number_str)} (row {row_idx})")
            elif unmatched_count == 21:
                print("⚠️ ... suppressing further 'No sites found' lines ...")
            continue

        item_no = row[item_col - 1].value if item_col else ''
        description = row[desc_col - 1].value if desc_col else ''
        quantity = row[qty_col - 1].value if qty_col else 1
        unit = row[unit_col - 1].value if unit_col else 'PCE'
        region_value = row[region_col - 1].value if region_col else None

        po_date = row[po_date_col - 1].value if po_date_col else None
        if isinstance(po_date, datetime):
            po_date = po_date.strftime('%Y-%m-%d')
        elif po_date:
            po_date = str(po_date)

        for site in sites:
            print(f"   ➜ Assigning PO {po_number_str} to SMP: {site.smp} - {site.site_name}")

            po_item, created = PurchaseOrder.objects.get_or_create(
                site=site,
                po_number=po_number_str,
                item_no=str(item_no).strip() if item_no else '',
                defaults={
                    'po_date': po_date or '',
                    'unit': str(unit).strip() if unit else 'PCE',
                    'quantity': float(quantity) if quantity else 1,
                    'description': str(description).strip(),
                    'source_file': uploaded_file
                }
            )

            if not created:
                po_item.po_date = po_date or ''
                po_item.unit = str(unit).strip() if unit else 'PCE'
                po_item.quantity = float(quantity) if quantity else 1
                po_item.description = str(description).strip()
                po_item.source_file = uploaded_file
                po_item.save()

            # Update region from PO file if available
            if region_value:
                region_str = str(region_value).strip()
                if region_str and site.region != region_str:
                    site.region = region_str
                    site.save(update_fields=['region'])
                    regions_corrected += 1
                    print(f"   ✅ Region updated for {site.smp}: {site.region}")

            sites_updated.add(site.id)
            po_count += 1

    try:
        wb.close()
    except:
        pass

    print(f"✅ Purchase Orders processed: {po_count} PO lines for {len(sites_updated)} sites "
          f"({unmatched_count} PO line(s) had no matching site)")
    if region_col:
        print(f"✅ Region accuracy: corrected {regions_corrected} site region value(s)")
    return po_count


def convert_excel_to_pdf_powershell(excel_path, pdf_path, _attempt=1, _max_attempts=3):
    """Convert Excel to PDF using PowerShell + Excel COM."""
    if not IS_WINDOWS:
        return False

    with EXCEL_COM_LOCK:
        try:
            kill_excel_processes()
            time.sleep(2)

            print("🔄 Converting Excel to PDF using PowerShell...")
            excel_path_escaped = excel_path.replace('\\', '\\\\')
            pdf_path_escaped = pdf_path.replace('\\', '\\\\')

            ps_script = f'''
            $ErrorActionPreference = "Stop"
            try {{
                $excel = New-Object -ComObject Excel.Application
                $excel.Visible = $false
                $excel.DisplayAlerts = $false
                $excel.AutomationSecurity = 3
                $excel.Interactive = $false
                $excel.EnableEvents = $false
                $workbook = $excel.Workbooks.Open("{excel_path_escaped}", $null, $null, $null, $null, $null, $true)
                $workbook.ExportAsFixedFormat(0, "{pdf_path_escaped}")
                $workbook.Close($false)
                $excel.Quit()
                [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
            }} catch {{
                Write-Error $_.Exception.Message
                exit 1
            }}
            '''

            result = subprocess.run(
                ['powershell', '-Command', ps_script],
                capture_output=True,
                text=True,
                timeout=90
            )

            if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 1000:
                print(f"✅ PDF saved via PowerShell: {pdf_path} ({os.path.getsize(pdf_path)} bytes)")
                return True

            combined_output = (result.stderr or '') + (result.stdout or '')
            if _is_call_rejected_error(combined_output) and _attempt < _max_attempts:
                wait = 3 * _attempt
                print(f"⚠️ RPC_E_CALL_REJECTED via PowerShell (attempt {_attempt}/{_max_attempts}) — retrying in {wait}s...")
                kill_excel_processes()
                time.sleep(wait)
                return convert_excel_to_pdf_powershell(excel_path, pdf_path, _attempt=_attempt + 1, _max_attempts=_max_attempts)

            print(f"⚠️ PowerShell conversion failed: {result.stderr}")
            return False
        except Exception as e:
            if _is_call_rejected_error(e) and _attempt < _max_attempts:
                wait = 3 * _attempt
                print(f"⚠️ RPC_E_CALL_REJECTED via PowerShell (attempt {_attempt}/{_max_attempts}) — retrying in {wait}s...")
                kill_excel_processes()
                time.sleep(wait)
                return convert_excel_to_pdf_powershell(excel_path, pdf_path, _attempt=_attempt + 1, _max_attempts=_max_attempts)
            print(f"⚠️ PowerShell method failed: {e}")
            return False


def convert_excel_to_pdf_libreoffice(excel_path, pdf_path):
    """Convert a single Excel file to PDF using LibreOffice headless."""
    try:
        print("🔄 Converting Excel to PDF using LibreOffice headless...")
        out_dir = os.path.dirname(pdf_path)
        os.makedirs(out_dir, exist_ok=True)

        kill_excel_processes()
        time.sleep(1)

        soffice_cmd = _find_soffice()
        if not soffice_cmd:
            print("⚠️ LibreOffice executable not found")
            return False

        print(f"🔄 Using LibreOffice at: {soffice_cmd}")

        result = subprocess.run(
            [
                soffice_cmd, 
                '--headless', 
                '--convert-to', 'pdf:calc_pdf_Export',
                '--outdir', out_dir, 
                excel_path
            ],
            capture_output=True,
            text=True,
            timeout=120
        )

        expected_name = os.path.splitext(os.path.basename(excel_path))[0] + '.pdf'
        produced_path = os.path.join(out_dir, expected_name)

        if os.path.exists(produced_path):
            if produced_path != pdf_path:
                shutil.move(produced_path, pdf_path)
            print(f"✅ PDF saved via LibreOffice: {pdf_path} ({os.path.getsize(pdf_path)} bytes)")
            return True

        print(f"⚠️ LibreOffice conversion failed: {result.stderr or result.stdout}")
        return False
    except FileNotFoundError:
        print("⚠️ LibreOffice ('soffice') not found on this server.")
        return False
    except Exception as e:
        print(f"⚠️ LibreOffice method failed: {e}")
        return False


def convert_excel_batch_to_pdf_libreoffice(excel_paths, pdf_dir, timeout_per_file=20, base_timeout=60):
    """Convert MANY Excel files to PDF in a single LibreOffice headless invocation."""
    results = {p: False for p in excel_paths}
    if not excel_paths:
        return results

    try:
        os.makedirs(pdf_dir, exist_ok=True)
        kill_excel_processes()
        time.sleep(1)

        soffice_cmd = _find_soffice()
        if not soffice_cmd:
            print("⚠️ LibreOffice executable not found — batch PDF conversion skipped")
            return results

        print(f"🔄 Batch-converting {len(excel_paths)} certificate(s) to PDF via LibreOffice...")

        timeout = base_timeout + timeout_per_file * len(excel_paths)
        
        result = subprocess.run(
            [
                soffice_cmd, 
                '--headless', 
                '--convert-to', 'pdf:calc_pdf_Export',
                '--outdir', pdf_dir, 
                *excel_paths
            ],
            capture_output=True,
            text=True,
            timeout=timeout
        )

        for excel_path in excel_paths:
            expected_name = os.path.splitext(os.path.basename(excel_path))[0] + '.pdf'
            produced_path = os.path.join(pdf_dir, expected_name)
            if os.path.exists(produced_path) and os.path.getsize(produced_path) > 1000:
                results[excel_path] = True

        succeeded = sum(1 for v in results.values() if v)
        print(f"✅ Batch PDF conversion: {succeeded}/{len(excel_paths)} succeeded")

        if succeeded < len(excel_paths):
            print(f"⚠️ Batch conversion output: {result.stderr or result.stdout}")

        return results

    except Exception as e:
        print(f"⚠️ Batch LibreOffice conversion failed: {e}")
        traceback.print_exc()
        return results


def convert_excel_to_pdf(excel_path, pdf_path):
    """Convert Excel to PDF using the best available method."""
    kill_excel_processes()
    time.sleep(2)

    if check_libreoffice_available():
        if convert_excel_to_pdf_libreoffice(excel_path, pdf_path):
            return True

    if IS_WINDOWS:
        ensure_excel_com_desktop_folder()
        if convert_excel_to_pdf_powershell(excel_path, pdf_path):
            return True

    print("⚠️ All PDF conversion methods failed")
    return False


def _pdf_dir_for(output_path):
    """Given the path of the single-sheet certificate .xlsx, return the matching .../certificates/pdf/ directory."""
    excel_dir = os.path.dirname(output_path)
    if f"{os.sep}excel" in excel_dir:
        pdf_dir = excel_dir.replace(f"{os.sep}excel", f"{os.sep}pdf")
    elif excel_dir.endswith('excel'):
        pdf_dir = excel_dir[: -len('excel')] + 'pdf'
    else:
        pdf_dir = os.path.join(os.path.dirname(excel_dir), 'pdf')
    os.makedirs(pdf_dir, exist_ok=True)
    return pdf_dir


ITEM_NO_COL = 2
DESCRIPTION_COL = 3
QTY_COL = 8
REMARKS_COL = 9


# ============================================================
# DATE FORMATTING
# ============================================================
_DATE_INPUT_FORMATS = (
    '%Y-%m-%d',
    '%Y/%m/%d',
    '%d-%m-%Y',
    '%d/%m/%Y',
    '%m-%d-%Y',
    '%m/%d/%Y',
    '%d %b %Y',
    '%d %B %Y',
    '%Y-%m-%d %H:%M:%S',
)


def format_display_date(value):
    """Format a date value as 'D Mon YYYY', e.g. '7 Aug 2026'."""
    if not value:
        return ''

    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value).strip()
        if not raw:
            return ''
        dt = None
        for fmt in _DATE_INPUT_FORMATS:
            try:
                dt = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            return raw

    return f"{dt.day} {dt.strftime('%b')} {dt.year}"


# ============================================================
# CERTIFICATE PAGE SIZING
# ============================================================
def _cm_to_in(value_cm):
    return round(value_cm / 2.54, 4)


CERTIFICATE_PAGE_MARGINS = PageMargins(
    top=_cm_to_in(1.91),
    bottom=_cm_to_in(1.91),
    left=_cm_to_in(1.87),
    right=_cm_to_in(1.87),
    header=_cm_to_in(0.6),
    footer=_cm_to_in(0.6),
)


def find_items_header_row(ws, search_rows=30):
    """Locate the items-table header row"""
    for row in range(1, search_rows + 1):
        cell_value = ws.cell(row=row, column=DESCRIPTION_COL).value
        if cell_value and 'SID' in str(cell_value):
            return row
    print("⚠️ Could not locate 'SID (Description)' header - defaulting to row 12")
    return 12


def find_items_table_end(ws, start_row, search_rows=60):
    """Find the row right BEFORE the 'Check Points' section"""
    for row in range(start_row, start_row + search_rows):
        for col in range(1, 10):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and 'Check Points' in str(cell_value):
                return row - 1
    print("⚠️ Could not locate 'Check Points' - falling back to a 20-row cap")
    return start_row + 19


def clear_items_table(ws, start_row, end_row, max_cols=10):
    """Completely clear all cells from start_row through end_row"""
    print(f"🗑️ CLEARING items table rows {start_row} to {end_row}")

    cleared_count = 0
    for row in range(start_row, end_row + 1):
        if row > ws.max_row:
            break
        for col in range(1, max_cols + 1):
            if not is_merged_write_blocked(ws, row, col):
                ws.cell(row=row, column=col).value = None
                cleared_count += 1

    print(f"   ✅ Cleared {cleared_count} cells")
    return True


def fill_items_table(ws, po_items, start_row, end_row):
    """Fill items table with PO data"""
    print(f"📝 Filling {po_items.count()} items from Purchase Orders...")

    filled_count = 0

    for idx, po_item in enumerate(po_items[:20]):
        row = start_row + idx
        if row > end_row:
            print(f"   ⚠️ Stopped at row {row}: would overrun into the Check Points section")
            break

        if not is_merged_write_blocked(ws, row, ITEM_NO_COL):
            ws.cell(row=row, column=ITEM_NO_COL).value = po_item.item_no if po_item.item_no else idx + 1

        # DESCRIPTION - ALWAYS write to the anchor cell of the merged range
        # The description column (C) is merged with D:G, so we must write to the anchor
        anchor_row, anchor_col = get_merged_anchor(ws, row, DESCRIPTION_COL)
        ws.cell(row=anchor_row, column=anchor_col).value = po_item.description

        if not is_merged_write_blocked(ws, row, QTY_COL):
            ws.cell(row=row, column=QTY_COL).value = float(po_item.quantity) if po_item.quantity else 1

        if getattr(po_item, 'remarks', None) and not is_merged_write_blocked(ws, row, REMARKS_COL):
            ws.cell(row=row, column=REMARKS_COL).value = po_item.remarks

        filled_count += 1

    print(f"   ✅ Filled {filled_count} items")
    return filled_count


def update_bottom_dates(ws, current_date):
    """Set the two signature-block dates to the current date"""
    date_row = None
    for row in range(1, 40):
        label = ws.cell(row=row, column=2).value
        if label and str(label).strip().lower().startswith('date'):
            date_row = row
            break

    if not date_row:
        print("⚠️ Could not locate the 'Date:' row - bottom dates not updated")
        return

    for col in (3, 7):
        if not is_merged_write_blocked(ws, date_row, col):
            ws.cell(row=date_row, column=col).value = current_date

    print(f"✅ Bottom dates (row {date_row}) updated to: {current_date}")


# ============================================================
# FIELD PLACEMENT
# ============================================================

def _find_label_cell(ws, label_variants, search_rows=40, search_cols=15):
    """Find the (row, col) of a cell whose text matches one of `label_variants`."""
    normalized_variants = [v.lower().strip().rstrip(':').strip() for v in label_variants]

    max_row = min(search_rows, ws.max_row)
    max_col = min(search_cols, ws.max_column)

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if not cell.value:
                continue
            text = str(cell.value).lower().strip().rstrip(':').strip()
            if text in normalized_variants:
                return cell.row, cell.column

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if not cell.value:
                continue
            text = str(cell.value).lower().strip().rstrip(':').strip()
            for variant in normalized_variants:
                if variant and (variant in text or text in variant):
                    return cell.row, cell.column

    return None, None


def _next_writable_cell_right(ws, row, col, max_search=6):
    """Walk right from (row, col) and return the first cell that isn't blocked by a merged range."""
    for c in range(col + 1, col + 1 + max_search):
        if not is_merged_write_blocked(ws, row, c):
            return row, c
    return row, col + 1


def _set_field_by_token_or_placeholder(ws, token, value, legacy_placeholders, field_label, results, written_cells, label_variants=None):
    """Fill a single field on the certificate."""
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in written_cells:
                continue
            if cell.value and str(cell.value).strip() == token:
                cell.value = value
                written_cells.add((cell.row, cell.column))
                results[field_label] = 'token'
                return True

    for placeholder in legacy_placeholders:
        for row in ws.iter_rows():
            for cell in row:
                if (cell.row, cell.column) in written_cells:
                    continue
                if cell.value and placeholder in str(cell.value):
                    cell.value = value
                    written_cells.add((cell.row, cell.column))
                    results[field_label] = f'legacy match ("{placeholder}")'
                    return True

    if label_variants:
        label_row, label_col = _find_label_cell(ws, label_variants)
        if label_row is not None:
            target_row, target_col = _next_writable_cell_right(ws, label_row, label_col)
            if (target_row, target_col) not in written_cells:
                ws.cell(row=target_row, column=target_col).value = value
                written_cells.add((target_row, target_col))
                results[field_label] = f'label match (row {label_row})'
                return True

    results[field_label] = 'NOT FOUND'
    return False


def update_book2_status(book2_path, smp, status='Completed'):
    """Update the WCC Status column in the Book2 file for a single SMP."""
    if not os.path.exists(book2_path):
        print(f"⚠️ Book2 file not found: {book2_path}")
        return False

    try:
        kill_excel_processes()
        time.sleep(1)

        wb = openpyxl.load_workbook(book2_path)
        ws = wb.active

        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        smp_col, _ = _find_header_col(headers, ['Smp', 'SMP', 'SMP ID'])
        wcc_col, _ = _find_header_col(headers, ['WCC Status', 'WCC_STATUS', 'Status'])

        if not smp_col or not wcc_col:
            print(f"⚠️ Could not find SMP or WCC Status columns in Book2")
            try:
                wb.close()
            except:
                pass
            return False

        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=smp_col).value
            if cell_value and str(cell_value).strip() == str(smp).strip():
                ws.cell(row=row_idx, column=wcc_col).value = status
                found = True
                print(f"✅ Updated WCC Status for SMP {smp} to '{status}' in Book2")
                break

        if not found:
            print(f"⚠️ SMP {smp} not found in Book2")

        success = safe_save_workbook(wb, book2_path)
        return success and found

    except Exception as e:
        print(f"❌ Error updating Book2 status: {e}")
        return False


def update_book2_status_bulk(book2_path, smp_list, status='Completed'):
    """Update WCC Status for MANY SMPs in a single load/save pass."""
    if not os.path.exists(book2_path):
        print(f"⚠️ Book2 file not found: {book2_path}")
        return False

    if not smp_list:
        return True

    smp_set = {str(s).strip() for s in smp_list if s}
    if not smp_set:
        return True

    try:
        kill_excel_processes()
        time.sleep(1)

        wb = openpyxl.load_workbook(book2_path)
        ws = wb.active

        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        smp_col, _ = _find_header_col(headers, ['Smp', 'SMP', 'SMP ID'])
        wcc_col, _ = _find_header_col(headers, ['WCC Status', 'WCC_STATUS', 'Status'])

        if not smp_col or not wcc_col:
            print("⚠️ Could not find SMP or WCC Status columns in Book2")
            try:
                wb.close()
            except:
                pass
            return False

        updated = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=smp_col).value
            if cell_value and str(cell_value).strip() in smp_set:
                ws.cell(row=row_idx, column=wcc_col).value = status
                updated += 1

        print(f"✅ Bulk-updated WCC Status to '{status}' for {updated}/{len(smp_set)} SMP(s) in Book2")

        return safe_save_workbook(wb, book2_path)

    except Exception as e:
        print(f"❌ Error bulk-updating Book2 status: {e}")
        traceback.print_exc()
        return False


def generate_sorted_excel(sites, output_path):
    """Generate a sorted Excel file with all sites and their PO items."""
    print(f"📂 Generating sorted Excel: {output_path}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sites Data"

    headers = [
        'No.', 'SMP', 'Site ID', 'Site Name', 'PO Number', 'PO Date',
        'Site Module Package', 'Project Category', 'Nokia PM', 'Region',
        'WCC Status', 'Item No', 'Item Description', 'Quantity', 'Unit'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="1e6b3a", end_color="1e6b3a", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    row_idx = 2
    for site in sites:
        po_items = PurchaseOrder.objects.filter(site=site)
        po_date_display = format_display_date(site.po_date)

        if po_items.exists():
            for po in po_items:
                ws.cell(row=row_idx, column=1).value = site.no or ''
                ws.cell(row=row_idx, column=2).value = site.smp
                ws.cell(row=row_idx, column=3).value = site.site_id
                ws.cell(row=row_idx, column=4).value = site.site_name
                ws.cell(row=row_idx, column=5).value = site.po_number
                ws.cell(row=row_idx, column=6).value = po_date_display
                ws.cell(row=row_idx, column=7).value = site.site_module_package
                ws.cell(row=row_idx, column=8).value = site.project_category
                ws.cell(row=row_idx, column=9).value = site.nokia_pm
                ws.cell(row=row_idx, column=10).value = site.region
                ws.cell(row=row_idx, column=11).value = site.wcc_status
                ws.cell(row=row_idx, column=12).value = po.item_no
                ws.cell(row=row_idx, column=13).value = po.description
                ws.cell(row=row_idx, column=14).value = po.quantity
                ws.cell(row=row_idx, column=15).value = po.unit
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1).value = site.no or ''
            ws.cell(row=row_idx, column=2).value = site.smp
            ws.cell(row=row_idx, column=3).value = site.site_id
            ws.cell(row=row_idx, column=4).value = site.site_name
            ws.cell(row=row_idx, column=5).value = site.po_number
            ws.cell(row=row_idx, column=6).value = po_date_display
            ws.cell(row=row_idx, column=7).value = site.site_module_package
            ws.cell(row=row_idx, column=8).value = site.project_category
            ws.cell(row=row_idx, column=9).value = site.nokia_pm
            ws.cell(row=row_idx, column=10).value = site.region
            ws.cell(row=row_idx, column=11).value = site.wcc_status
            row_idx += 1

    for col in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, min(row_idx, 100) + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    safe_save_workbook(wb, output_path)
    print(f"✅ Sorted Excel saved: {output_path}")
    return output_path


def generate_certificate_excel(site, po_items, template_path, output_path, convert_pdf=True):
    """Generate certificate for a SINGLE site with its PO items."""
    print(f"🔍 Generating certificate for site: {site.smp} - {site.site_name}")
    print(f"📋 PO Number: {site.po_number}")
    print(f"📂 Site Module Package: {site.site_module_package}")
    print(f"📄 Template: {template_path}")
    print(f"📁 Output: {output_path}")

    try:
        kill_excel_processes()
        time.sleep(1)

        sheet_name = get_sheet_from_module(site.site_module_package)
        print(f"📋 Target sheet: {sheet_name}")

        success = copy_sheet_with_openpyxl(template_path, sheet_name, output_path)

        if not success:
            raise Exception(f"Failed to create workbook: {output_path}")

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # ============================================================
        # CERTIFICATE PAGE SIZING
        # ============================================================
        try:
            ws.page_setup.paperSize = 1  # Letter
            ws.page_setup.orientation = 'landscape'
            
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.scale = None
            
            if ws.dimensions and ws.dimensions != 'A1:A1':
                ws.print_area = f"A1:{ws.dimensions.split(':')[1]}"
            
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = True
            
            def cm_to_in(cm):
                return round(cm / 2.54, 4)
            
            ws.page_margins = PageMargins(
                top=cm_to_in(1.91),
                bottom=cm_to_in(1.91),
                left=cm_to_in(1.87),
                right=cm_to_in(1.87),
                header=cm_to_in(0.6),
                footer=cm_to_in(0.6)
            )
            
            print("📐 Applied certificate sizing: Letter · margins T/B 1.91cm, L/R 1.87cm · fitToPage=1 page")
            
        except Exception as e:
            print(f"⚠️ Could not apply certificate page sizing: {e}")
            traceback.print_exc()

        current_date = format_display_date(datetime.now())

        print("📝 Filling site data from Book2...")
        field_results = {}
        written_cells = set()

        # Site ID & Site Name: "{Site ID}-{Site Name}"
        site_display = f"{site.site_id}-{site.site_name}"
        _set_field_by_token_or_placeholder(
            ws, '{{SITE_ID}}', site_display,
            [
                "MOS0201-Malla Shopping Centre",
                "BIA0073-Busia_Amoni",
                "VGA0038-Luanda",
                "MOIBEN-UGU0053",
                "KGA0003-NYI0004-IHWAGI-KARATINA",
                "TKA0129-TKA0009-Kihingo Hill-GATUNDU_NORTH"
            ],
            'Site ID & Name', field_results, written_cells,
            label_variants=['Site ID & Site Name', 'Site ID and Site Name', 'Site ID & Name']
        )

        region_value = site.region or 'Unknown'
        _set_field_by_token_or_placeholder(
            ws, '{{REGION}}', region_value,
            ["Machakos", "Busia", "VIHIGA", "UASIN GISHU", "Nyeri", "Thika"],
            'Region', field_results, written_cells,
            label_variants=['Region']
        )

        site_type = 'Green Field'
        if site.sub_category == 'Modernization':
            site_type = 'Modernization'
        elif site.sub_category in ['TSS', 'Los', 'Survey']:
            site_type = 'Greenfield'
        _set_field_by_token_or_placeholder(
            ws, '{{SITE_TYPE}}', site_type,
            ["Green Field", "Greenfield"],
            'Site Type', field_results, written_cells,
            label_variants=['Site Type']
        )

        smp_value = site.smp or ''
        _set_field_by_token_or_placeholder(
            ws, '{{SMP_ID}}', smp_value,
            [
                "SMP-WO-1724355", "SMP-WO-1851398", "SMP-WO-1851172",
                "SMP-WO-1395549", "SMP-WO-1827288", "SMP-WO-1719483"
            ],
            'SMP ID', field_results, written_cells,
            label_variants=['SMP ID', 'SMP']
        )

        po_number_value = site.po_number
        _set_field_by_token_or_placeholder(
            ws, '{{PO_NUMBER}}', po_number_value,
            ["51115097", "51977371", "51961036", "49325234", "51741122", "51693513"],
            'PO Number', field_results, written_cells,
            label_variants=['PO Number']
        )

        po_date_value = format_display_date(site.po_date)
        _set_field_by_token_or_placeholder(
            ws, '{{PO_DATE}}', po_date_value,
            ["2025-12-04", "2026-08-07", "2026-08-03", "2024-07-03", "2026-06-11", "2026-05-28"],
            'PO Date', field_results, written_cells,
            label_variants=['PO Date']
        )

        pm_value = site.nokia_pm or 'Julius Kamemba'
        _set_field_by_token_or_placeholder(
            ws, '{{PM_NAME}}', pm_value,
            ["Julius Kamemba", "Ken Juma"],
            'Nokia PM', field_results, written_cells,
            label_variants=['Nokia Project Manager/ROM', 'Nokia Project Manager', 'Project Manager']
        )

        print("📊 Field mapping summary:")
        for label, outcome in field_results.items():
            icon = '✅' if outcome != 'NOT FOUND' else '❌'
            print(f"   {icon} {label}: {outcome}")

        print("📝 Processing items table...")

        items_header_row = find_items_header_row(ws)
        items_start_row = items_header_row + 1
        items_end_row = find_items_table_end(ws, items_start_row)

        print(f"📊 Items table: header row {items_header_row}, data rows {items_start_row}-{items_end_row}")

        if items_start_row and po_items.exists():
            clear_items_table(ws, items_start_row, items_end_row, max_cols=10)
            fill_items_table(ws, po_items, items_start_row, items_end_row)
        else:
            print("⚠️ No PO items to write")

        update_bottom_dates(ws, current_date)

        success = safe_save_workbook(wb, output_path)

        if not success:
            raise Exception(f"Failed to save workbook after multiple attempts: {output_path}")

        print(f"✅ Excel saved: {output_path}")

        if not convert_pdf:
            return output_path, None

        base_name = os.path.splitext(os.path.basename(output_path))[0]
        pdf_dir = _pdf_dir_for(output_path)
        signed_pdf_path = os.path.join(pdf_dir, f"{base_name}.pdf")

        kill_excel_processes()
        time.sleep(2)

        pdf_created = convert_excel_to_pdf(output_path, signed_pdf_path)

        if pdf_created and os.path.exists(signed_pdf_path) and os.path.getsize(signed_pdf_path) > 1000:
            print(f"✅ PDF with signatures created: {signed_pdf_path}")
        else:
            print("⚠️ PDF with signatures could not be created")
            signed_pdf_path = None

        return output_path, signed_pdf_path

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        traceback.print_exc()
        raise e